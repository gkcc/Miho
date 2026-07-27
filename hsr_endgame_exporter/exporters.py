from __future__ import annotations

import csv
import importlib.util
from collections import defaultdict
from pathlib import Path
from typing import Any

from .constants import MODE_CN
from .constants import (
    CHARACTER_USAGE_COLUMNS,
    HISTOGRAPH_COLUMNS,
    NAME_MAP_COLUMNS,
    PHASE_COLUMNS,
    PRYDWEN_TIER_CHANGELOG_COLUMNS,
    PRYDWEN_TIER_CHART_COLUMNS,
    PRYDWEN_TIER_COLUMNS,
    PRYDWEN_TIER_USAGE_TREND_COLUMNS,
    TEAM_ORDERED_COLUMNS,
    TEAM_RAW_COLUMNS,
    TEAM_UNORDERED_COLUMNS,
)
from .normalize import natural_version_key
from .parsers import attach_team_signatures

DISPLAY_TOP_TEAMS_PER_MODE = 100


def write_csv(path: Path, rows: list[dict[str, Any]], columns: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({column: _csv_value(row.get(column)) for column in columns})


def latest_character_usage(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    chosen: dict[tuple[str, str, str, str], dict[str, Any]] = {}
    for row in rows:
        key = (
            row.get("mode", ""),
            row.get("sub_mode", ""),
            row.get("phase_ver", ""),
            row.get("character_slug", ""),
        )
        current = chosen.get(key)
        if current is None or str(row.get("collect_date", "")) >= str(current.get("collect_date", "")):
            chosen[key] = row
    return sorted(
        chosen.values(),
        key=lambda row: (
            row.get("mode", ""),
            row.get("sub_mode", ""),
            row.get("phase_ver", ""),
            row.get("character_slug", ""),
        ),
    )


def dedup_ordered_teams(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        ordered, _ = attach_team_signatures(row)
        groups[ordered].append(row)

    output: list[dict[str, Any]] = []
    for signature, group in groups.items():
        best = dict(sorted(group, key=_team_sort_key)[0])
        best["ordered_signature"] = signature
        best["duplicate_count"] = len(group)
        best["merged_source_files"] = ";".join(
            sorted({str(item.get("source_file", "")) for item in group if item.get("source_file")})
        )
        output.append(best)
    return sorted(output, key=_team_output_key)


def dedup_unordered_teams(ordered_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in ordered_rows:
        _, unordered = attach_team_signatures(row)
        groups[unordered].append(row)

    output: list[dict[str, Any]] = []
    for signature, group in groups.items():
        best = dict(sorted(group, key=_team_sort_key)[0])
        best["unordered_signature"] = signature
        best["ordered_signature_examples"] = ";".join(
            sorted({str(item.get("ordered_signature", "")) for item in group if item.get("ordered_signature")})
        )
        best["duplicate_count"] = sum(int(item.get("duplicate_count") or 1) for item in group)
        output.append(best)
    return sorted(output, key=_team_output_key)


def write_all_outputs(
    out_dir: Path,
    *,
    phase_rows: list[dict[str, Any]],
    character_rows: list[dict[str, Any]],
    histograph_rows: list[dict[str, Any]],
    team_raw_rows: list[dict[str, Any]],
    name_map_rows: list[dict[str, Any]],
    name_map_unresolved_rows: list[dict[str, Any]],
    prydwen_tier_current_rows: list[dict[str, Any]] | None = None,
    prydwen_tier_history_rows: list[dict[str, Any]] | None = None,
    prydwen_tier_changelog_rows: list[dict[str, Any]] | None = None,
    prydwen_tier_changelog_history_rows: list[dict[str, Any]] | None = None,
    prydwen_tier_usage_trend_rows: list[dict[str, Any]] | None = None,
    prydwen_tier_chart_rows: list[dict[str, Any]] | None = None,
    warnings: list[str],
) -> dict[str, list[dict[str, Any]]]:
    latest_rows = latest_character_usage(character_rows)
    team_ordered_rows = dedup_ordered_teams(team_raw_rows)
    team_unordered_rows = dedup_unordered_teams(team_ordered_rows)

    tables = {
        "phase_index": phase_rows,
        "character_usage_long": character_rows,
        "character_usage_phase_latest": latest_rows,
        "histograph_usage_long": histograph_rows,
        "team_rank_raw": team_raw_rows,
        "team_rank_dedup_ordered": team_ordered_rows,
        "team_rank_dedup_unordered": team_unordered_rows,
        "name_map": name_map_rows,
        "name_map_unresolved": name_map_unresolved_rows,
        "prydwen_tier_current": prydwen_tier_current_rows or [],
        "prydwen_tier_history": prydwen_tier_history_rows or [],
        "prydwen_tier_changelog": prydwen_tier_changelog_rows or [],
        "prydwen_tier_changelog_history": prydwen_tier_changelog_history_rows or [],
        "prydwen_tier_usage_trend": prydwen_tier_usage_trend_rows or [],
        "prydwen_tier_charts": prydwen_tier_chart_rows or [],
    }
    overview_rows = _build_overview_rows(tables, warnings)
    latest_usage_cn_rows = _build_latest_usage_cn(tables)
    top_teams_latest_rows = _build_top_teams_latest(tables)
    write_csv(out_dir / "phase_index.csv", phase_rows, PHASE_COLUMNS)
    write_csv(out_dir / "character_usage_long.csv", character_rows, CHARACTER_USAGE_COLUMNS)
    write_csv(out_dir / "character_usage_phase_latest.csv", latest_rows, CHARACTER_USAGE_COLUMNS)
    write_csv(out_dir / "histograph_usage_long.csv", histograph_rows, HISTOGRAPH_COLUMNS)
    write_csv(out_dir / "team_rank_raw.csv", team_raw_rows, TEAM_RAW_COLUMNS)
    write_csv(out_dir / "team_rank_dedup_ordered.csv", team_ordered_rows, TEAM_ORDERED_COLUMNS)
    write_csv(out_dir / "team_rank_dedup_unordered.csv", team_unordered_rows, TEAM_UNORDERED_COLUMNS)
    write_csv(out_dir / "name_map.csv", name_map_rows, NAME_MAP_COLUMNS)
    write_csv(out_dir / "name_map_unresolved.csv", name_map_unresolved_rows, NAME_MAP_COLUMNS)
    write_csv(out_dir / "prydwen_tier_current.csv", tables["prydwen_tier_current"], PRYDWEN_TIER_COLUMNS)
    write_csv(out_dir / "prydwen_tier_history.csv", tables["prydwen_tier_history"], PRYDWEN_TIER_COLUMNS)
    write_csv(out_dir / "prydwen_tier_changelog.csv", tables["prydwen_tier_changelog"], PRYDWEN_TIER_CHANGELOG_COLUMNS)
    write_csv(out_dir / "prydwen_tier_changelog_history.csv", tables["prydwen_tier_changelog_history"], PRYDWEN_TIER_CHANGELOG_COLUMNS)
    write_csv(out_dir / "prydwen_tier_usage_trend.csv", tables["prydwen_tier_usage_trend"], PRYDWEN_TIER_USAGE_TREND_COLUMNS)
    write_csv(out_dir / "prydwen_tier_charts.csv", tables["prydwen_tier_charts"], PRYDWEN_TIER_CHART_COLUMNS)
    write_csv(out_dir / "overview.csv", overview_rows, ["section", "metric", "value"])
    write_csv(out_dir / "latest_usage_cn.csv", latest_usage_cn_rows, _columns_from_rows(latest_usage_cn_rows))
    write_csv(out_dir / "top_teams_latest.csv", top_teams_latest_rows, _columns_from_rows(top_teams_latest_rows))
    write_excel_if_available(out_dir / "hsr_endgame_dataset.xlsx", tables, warnings)
    return tables


def write_excel_if_available(
    path: Path,
    tables: dict[str, list[dict[str, Any]]],
    warnings: list[str],
) -> None:
    if not importlib.util.find_spec("pandas") or not importlib.util.find_spec("openpyxl"):
        warnings.append("XLSX skipped because pandas and openpyxl are not both installed")
        return
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    excel_tables = {
        "overview": _build_overview_rows(tables, warnings),
        "latest_usage_cn": _build_latest_usage_cn(tables),
        "top_teams_latest": _build_top_teams_latest(tables),
        **tables,
    }
    excel_columns = {
        "overview": ["section", "metric", "value"],
        "latest_usage_cn": _columns_from_rows(excel_tables["latest_usage_cn"]),
        "top_teams_latest": _columns_from_rows(excel_tables["top_teams_latest"]),
        "phase_index": PHASE_COLUMNS,
        "character_usage_long": CHARACTER_USAGE_COLUMNS,
        "character_usage_phase_latest": CHARACTER_USAGE_COLUMNS,
        "histograph_usage_long": HISTOGRAPH_COLUMNS,
        "team_rank_raw": TEAM_RAW_COLUMNS,
        "team_rank_dedup_ordered": TEAM_ORDERED_COLUMNS,
        "team_rank_dedup_unordered": TEAM_UNORDERED_COLUMNS,
        "name_map": NAME_MAP_COLUMNS,
        "name_map_unresolved": NAME_MAP_COLUMNS,
        "prydwen_tier_current": PRYDWEN_TIER_COLUMNS,
        "prydwen_tier_history": PRYDWEN_TIER_COLUMNS,
        "prydwen_tier_changelog": PRYDWEN_TIER_CHANGELOG_COLUMNS,
        "prydwen_tier_changelog_history": PRYDWEN_TIER_CHANGELOG_COLUMNS,
        "prydwen_tier_usage_trend": PRYDWEN_TIER_USAGE_TREND_COLUMNS,
        "prydwen_tier_charts": PRYDWEN_TIER_CHART_COLUMNS,
    }
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, rows in excel_tables.items():
            pd.DataFrame(rows, columns=excel_columns.get(sheet_name)).to_excel(
                writer,
                sheet_name=sheet_name[:31],
                index=False,
            )
    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="263238")
    header_font = Font(color="FFFFFF", bold=True)
    soft_fill = PatternFill("solid", fgColor="E8F3F1")
    for index, worksheet in enumerate(workbook.worksheets):
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.sheet_view.showGridLines = False
        if index == 0:
            worksheet.freeze_panes = "A1"
        for cell in worksheet[1]:
            cell.fill = soft_fill if worksheet.title in {"overview", "latest_usage_cn", "top_teams_latest"} else header_fill
            cell.font = Font(bold=True, color="1F2933") if worksheet.title in {"overview", "latest_usage_cn", "top_teams_latest"} else header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        _fit_columns(worksheet, get_column_letter)
        _format_numbers(worksheet)
    workbook.active = 0
    workbook.save(path)


def _csv_value(value: Any) -> Any:
    return "" if value is None else value


def _columns_from_rows(rows: list[dict[str, Any]]) -> list[str]:
    columns: list[str] = []
    for row in rows:
        for key in row:
            if key not in columns:
                columns.append(key)
    return columns


def _team_sort_key(row: dict[str, Any]) -> tuple[int, float, float]:
    source_priority = {
        "prydwen_page": 0,
        "hf_comps": 1,
    }.get(str(row.get("source_kind", "")), 2)
    rank = row.get("rank")
    rank_value = float(rank) if isinstance(rank, (int, float)) else 1_000_000.0
    app_rate = row.get("app_rate")
    app_value = float(app_rate) if isinstance(app_rate, (int, float)) else -1.0
    return source_priority, rank_value, -app_value


def _team_output_key(row: dict[str, Any]) -> tuple[str, str, str, float]:
    rank = row.get("rank")
    rank_value = float(rank) if isinstance(rank, (int, float)) else 1_000_000.0
    return (
        str(row.get("mode", "")),
        str(row.get("sub_mode", "")),
        str(row.get("phase_ver", "")),
        rank_value,
    )


def _build_overview_rows(
    tables: dict[str, list[dict[str, Any]]],
    warnings: list[str],
) -> list[dict[str, Any]]:
    phase_rows = tables.get("phase_index", [])
    character_rows = tables.get("character_usage_long", [])
    team_rows = tables.get("team_rank_raw", [])
    team_ordered_rows = tables.get("team_rank_dedup_ordered", [])
    team_unordered_rows = tables.get("team_rank_dedup_unordered", [])
    name_rows = tables.get("name_map", [])
    unresolved_rows = tables.get("name_map_unresolved", [])
    tier_rows = tables.get("prydwen_tier_current", [])
    tier_trend_rows = tables.get("prydwen_tier_usage_trend", [])
    tier_chart_rows = tables.get("prydwen_tier_charts", [])
    modes = sorted({row.get("mode") for row in phase_rows if row.get("mode")})
    rows = [
        {"section": "summary", "metric": "snapshots", "value": len({row.get("snapshot_id") for row in phase_rows})},
        {"section": "summary", "metric": "modes", "value": ", ".join(f"{MODE_CN.get(mode, mode)}({mode})" for mode in modes)},
        {"section": "rows", "metric": "character_usage_long", "value": len(character_rows)},
        {"section": "rows", "metric": "team_rank_raw", "value": len(team_rows)},
        {"section": "rows", "metric": "team_rank_dedup_ordered", "value": len(team_ordered_rows)},
        {"section": "rows", "metric": "team_rank_dedup_unordered", "value": len(team_unordered_rows)},
        {"section": "dedup", "metric": "raw_rows_removed_by_ordered_dedup", "value": len(team_rows) - len(team_ordered_rows)},
        {"section": "dedup", "metric": "raw_rows_removed_by_unordered_dedup", "value": len(team_rows) - len(team_unordered_rows)},
        {
            "section": "display",
            "metric": "top_teams_latest",
            "value": f"unordered unique Top {DISPLAY_TOP_TEAMS_PER_MODE} per mode",
        },
        {"section": "names", "metric": "name_map", "value": len(name_rows)},
        {"section": "names", "metric": "name_map_unresolved", "value": len(unresolved_rows)},
        {"section": "prydwen_tier", "metric": "current_rows", "value": len(tier_rows)},
        {"section": "prydwen_tier", "metric": "usage_trend_rows_t0_t2", "value": len(tier_trend_rows)},
        {"section": "prydwen_tier", "metric": "charts", "value": len(tier_chart_rows)},
        {"section": "quality", "metric": "warnings", "value": len(warnings)},
    ]
    for mode in modes:
        rows.append(
            {
                "section": "coverage",
                "metric": MODE_CN.get(mode, mode),
                "value": sum(1 for row in phase_rows if row.get("mode") == mode),
            }
        )
    for warning in warnings[:8]:
        rows.append({"section": "warning", "metric": "", "value": warning})
    return rows


def _build_latest_usage_cn(tables: dict[str, list[dict[str, Any]]]) -> list[dict[str, Any]]:
    rows = [
        row
        for row in tables.get("character_usage_long", [])
        if row.get("sub_mode") in {"all", "all_bosses"}
    ]
    latest_collect = {
        mode: max(str(row.get("collect_date", "")) for row in rows if row.get("mode") == mode)
        for mode in {row.get("mode") for row in rows}
    }
    by_character: dict[str, dict[str, Any]] = {}
    for row in rows:
        mode = row.get("mode")
        if str(row.get("collect_date", "")) != latest_collect.get(mode):
            continue
        slug = row.get("character_slug", "")
        target = by_character.setdefault(
            slug,
            {
                "character_name_cn": row.get("character_name_cn", ""),
                "character_name_en": row.get("character_name_en", ""),
                "character_slug": slug,
                "role": row.get("role", ""),
            },
        )
        target[f"{mode}_app_rate"] = row.get("app_rate")
        target[f"{mode}_avg_round"] = row.get("avg_round")
    output = list(by_character.values())
    for row in output:
        values = [
            float(row.get(f"{mode}_app_rate") or 0)
            for mode in ("moc", "pf", "as", "aa")
        ]
        row["max_app_rate"] = max(values) if values else 0
    return sorted(output, key=lambda row: row.get("max_app_rate", 0), reverse=True)


def _build_top_teams_latest(tables: dict[str, list[dict[str, Any]]]) -> list[dict[str, Any]]:
    rows = [
        row
        for row in tables.get("team_rank_dedup_unordered", [])
        if _is_comprehensive_scope(row.get("scope")) and row.get("rank") not in {"", None}
    ]
    latest_observation = {
        mode: max(
            (_top_team_observation(row) for row in rows if row.get("mode") == mode),
            key=_top_team_observation_sort_key,
        )
        for mode in {row.get("mode") for row in rows}
        if mode
    }
    grouped: dict[tuple[str, tuple[str, ...]], list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        mode = str(row.get("mode") or "")
        if _top_team_observation(row) != latest_observation.get(mode):
            continue
        chars = tuple(sorted(str(row.get(f"char_{index}_slug") or "") for index in range(1, 5)))
        grouped[(mode, chars)].append(row)

    selected: list[tuple[dict[str, Any], dict[str, Any]]] = []
    for group in grouped.values():
        representative = min(
            group,
            key=lambda row: (_team_sort_key(row), str(row.get("unordered_signature") or "")),
        )
        merged = dict(representative)
        merged["duplicate_count"] = sum(_team_duplicate_count(row.get("duplicate_count")) for row in group)
        merged["source_kind"] = _merged_team_values(row.get("source_kind") for row in group)
        merged["merged_source_files"] = _merged_team_values(
            row.get("merged_source_files") or row.get("source_file") for row in group
        )
        selected.append((representative, merged))

    output: list[dict[str, Any]] = []
    seen_per_mode: dict[str, int] = defaultdict(int)
    for _, row in sorted(
        selected,
        key=lambda pair: (_team_sort_key(pair[0]), str(pair[0].get("unordered_signature") or "")),
    ):
        mode = str(row.get("mode") or "")
        if seen_per_mode[mode] >= DISPLAY_TOP_TEAMS_PER_MODE:
            continue
        seen_per_mode[mode] += 1
        cn_names = [row.get(f"char_{index}_name_cn") or row.get(f"char_{index}_slug") for index in range(1, 5)]
        output.append(
            {
                "mode_cn": row.get("mode_cn", ""),
                "mode": mode,
                "sub_mode_cn": row.get("sub_mode_cn", ""),
                "sub_mode": row.get("sub_mode", ""),
                "phase_ver": row.get("phase_ver", ""),
                "rank": row.get("rank", ""),
                "team_cn": " / ".join(cn_names),
                "app_rate": row.get("app_rate", ""),
                "avg_round": row.get("avg_round", ""),
                "source_kind": row.get("source_kind", ""),
                "duplicate_count": row.get("duplicate_count", ""),
                "unordered_signature": row.get("unordered_signature", ""),
            }
        )
    return output


def _top_team_observation(row: dict[str, Any]) -> tuple[str, str, str]:
    return tuple(
        str(row.get(key) or "")
        for key in ("collect_date", "snapshot_id", "phase_ver")
    )


def _top_team_observation_sort_key(identity: tuple[str, str, str]) -> tuple[Any, ...]:
    collect_date, snapshot_id, phase_ver = identity
    return collect_date, natural_version_key(snapshot_id), natural_version_key(phase_ver)


def _is_comprehensive_scope(value: Any) -> bool:
    return str(value or "").strip().lower().replace("_", "-") in {"", "all", "top", "all-bosses"}


def _team_duplicate_count(value: Any) -> int:
    try:
        return max(1, int(str(value).strip()))
    except (TypeError, ValueError):
        return 1


def _merged_team_values(values: Any) -> str:
    return ";".join(
        sorted(
            {
                item.strip()
                for value in values
                for item in str(value or "").split(";")
                if item.strip()
            }
        )
    )


def _fit_columns(worksheet, get_column_letter) -> None:
    width_overrides = {
        "raw_json": 18,
        "source_url": 24,
        "source_file": 28,
        "merged_source_files": 28,
        "ordered_signature_examples": 28,
        "team_cn": 42,
    }
    headers = [cell.value for cell in worksheet[1]]
    sample_limit = min(worksheet.max_row, 250)
    for index, header in enumerate(headers, start=1):
        values = [str(header or "")]
        for row in range(2, sample_limit + 1):
            value = worksheet.cell(row=row, column=index).value
            if value is not None:
                values.append(str(value))
        width = min(max(len(value) for value in values) + 2, 36)
        if header in width_overrides:
            width = width_overrides[header]
        worksheet.column_dimensions[get_column_letter(index)].width = max(width, 8)


def _format_numbers(worksheet) -> None:
    headers = [cell.value for cell in worksheet[1]]
    for index, header in enumerate(headers, start=1):
        if header and ("app_rate" in str(header) or str(header) in {"max_app_rate"}):
            for cell in worksheet.iter_cols(min_col=index, max_col=index, min_row=2):
                for item in cell:
                    item.number_format = "0.00"
        elif header and ("avg_round" in str(header) or "sample" in str(header) or "rank" == str(header)):
            for cell in worksheet.iter_cols(min_col=index, max_col=index, min_row=2):
                for item in cell:
                    item.number_format = "0.00"
