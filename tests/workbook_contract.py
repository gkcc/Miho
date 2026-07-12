from __future__ import annotations

from datetime import date, datetime, time
from enum import Enum
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Color, Font, PatternFill
from openpyxl.utils import get_column_letter


class FormulaPolicy(str, Enum):
    """How formula cells are represented by the semantic comparator.

    The Python exporters currently let pandas turn any external text beginning
    with ``=`` into a formula.  ``PRESERVE`` freezes that legacy behaviour.
    ``EXTERNAL_TEXT`` is the approved safety normalization for the Rust writer:
    it treats every formula in these data-only workbooks as literal text.
    """

    PRESERVE = "preserve"
    EXTERNAL_TEXT = "external_text"


def workbook_semantics(
    path: Path,
    *,
    formula_policy: FormulaPolicy = FormulaPolicy.PRESERVE,
) -> dict[str, Any]:
    """Return stable workbook semantics without ZIP metadata or style IDs."""

    workbook = load_workbook(path, data_only=False, read_only=False)
    try:
        sheets = [
            _worksheet_semantics(worksheet, formula_policy=formula_policy)
            for worksheet in workbook.worksheets
        ]
        return {
            "file_name": path.name,
            "sheet_order": list(workbook.sheetnames),
            "active_sheet": workbook.active.title,
            "sheets": sheets,
        }
    finally:
        workbook.close()


def compare_workbooks(
    expected: Path,
    actual: Path,
    *,
    formula_policy: FormulaPolicy = FormulaPolicy.PRESERVE,
) -> list[str]:
    """Return human-readable semantic differences between two workbooks."""

    expected_semantics = workbook_semantics(expected, formula_policy=formula_policy)
    actual_semantics = workbook_semantics(actual, formula_policy=formula_policy)
    differences: list[str] = []
    _collect_differences(expected_semantics, actual_semantics, "$", differences)
    return differences


def assert_workbooks_equal(
    expected: Path,
    actual: Path,
    *,
    formula_policy: FormulaPolicy = FormulaPolicy.PRESERVE,
) -> None:
    """Assert semantic equality and show a bounded, path-oriented diff."""

    differences = compare_workbooks(
        expected,
        actual,
        formula_policy=formula_policy,
    )
    if differences:
        preview = "\n".join(f"- {item}" for item in differences[:40])
        remaining = len(differences) - 40
        suffix = f"\n- ... and {remaining} more difference(s)" if remaining > 0 else ""
        raise AssertionError(
            f"workbook semantics differ ({len(differences)} difference(s)):\n"
            f"{preview}{suffix}"
        )


def _worksheet_semantics(
    worksheet: Any,
    *,
    formula_policy: FormulaPolicy,
) -> dict[str, Any]:
    cells: list[dict[str, Any]] = []
    formula_cells: list[str] = []
    for row in worksheet.iter_rows(
        min_row=worksheet.min_row,
        max_row=worksheet.max_row,
        min_col=worksheet.min_column,
        max_col=worksheet.max_column,
    ):
        for cell in row:
            data_type = cell.data_type
            if data_type == "f":
                formula_cells.append(cell.coordinate)
                if formula_policy is FormulaPolicy.EXTERNAL_TEXT:
                    data_type = "s"
            cells.append(
                {
                    "coordinate": cell.coordinate,
                    "value": _cell_value(cell.value),
                    "data_type": data_type,
                    "number_format": cell.number_format,
                }
            )

    if formula_policy is FormulaPolicy.EXTERNAL_TEXT:
        formula_cells = []

    return {
        "title": worksheet.title,
        "dimensions": worksheet.calculate_dimension(),
        "headers": [_cell_value(cell.value) for cell in worksheet[1]],
        "cells": cells,
        "formula_count": len(formula_cells),
        "formula_cells": formula_cells,
        "freeze_panes": _freeze_panes(worksheet.freeze_panes),
        "auto_filter": worksheet.auto_filter.ref,
        "show_grid_lines": (
            True
            if worksheet.sheet_view.showGridLines is None
            else worksheet.sheet_view.showGridLines
        ),
        "header_styles": [_header_style(cell) for cell in worksheet[1]],
        "column_widths": {
            get_column_letter(index): _column_width(worksheet, index)
            for index in range(worksheet.min_column, worksheet.max_column + 1)
        },
    }


def _cell_value(value: Any) -> Any:
    if isinstance(value, (datetime, date, time)):
        return {"temporal_type": type(value).__name__, "value": value.isoformat()}
    return value


def _freeze_panes(value: Any) -> str | None:
    if value is None:
        return None
    return value.coordinate if hasattr(value, "coordinate") else str(value)


def _column_width(worksheet: Any, index: int) -> float | None:
    dimension = worksheet.column_dimensions.get(get_column_letter(index))
    if dimension is not None:
        return dimension.width
    # OOXML can encode adjacent equal-width columns as one min/max range.
    # openpyxl indexes only the first column of that range.
    for candidate in worksheet.column_dimensions.values():
        if (
            candidate.min is not None
            and candidate.max is not None
            and candidate.min <= index <= candidate.max
        ):
            return candidate.width
    return None


def _header_style(cell: Any) -> dict[str, Any]:
    return {
        "coordinate": cell.coordinate,
        "fill": _fill(cell.fill),
        "font": _font(cell.font),
        "alignment": _alignment(cell.alignment),
        "border": _border(cell.border),
    }


def _fill(fill: PatternFill) -> dict[str, Any]:
    return {
        "fill_type": fill.fill_type,
        "foreground": _color(fill.fgColor),
        # A solid fill only renders its foreground. openpyxl and
        # rust_xlsxwriter encode the unused background differently.
        "background": None if fill.fill_type == "solid" else _color(fill.bgColor),
    }


def _font(font: Font) -> dict[str, Any]:
    color = _color(font.color)
    if color is not None and color["type"] == "theme" and color["value"] == 1:
        color = None
    return {
        # Python's custom Font omits inherited defaults while rust_xlsxwriter
        # writes them explicitly. Normalize both to this workbook's defaults.
        "name": font.name or "Calibri",
        "size": font.sz if font.sz is not None else 11.0,
        "bold": font.b,
        "italic": font.i,
        "underline": font.u,
        "strike": font.strike,
        "color": color,
        "vertical_alignment": font.vertAlign,
        "family": font.family if font.family is not None else 2.0,
        "scheme": font.scheme or "minor",
    }


def _border(border: Any) -> dict[str, Any]:
    return {
        "left": border.left.style,
        "right": border.right.style,
        "top": border.top.style,
        "bottom": border.bottom.style,
        "diagonal_up": border.diagonalUp,
        "diagonal_down": border.diagonalDown,
    }


def _alignment(alignment: Alignment) -> dict[str, Any]:
    return {
        "horizontal": alignment.horizontal,
        "vertical": alignment.vertical,
        "text_rotation": alignment.textRotation,
        "wrap_text": alignment.wrapText,
        "shrink_to_fit": alignment.shrinkToFit,
        "indent": alignment.indent,
    }


def _color(color: Color | None) -> dict[str, Any] | None:
    if color is None:
        return None
    value: str | int | None
    if color.type == "rgb":
        value = color.rgb[-6:] if color.rgb is not None else None
    elif color.type == "indexed":
        value = color.indexed
    elif color.type == "theme":
        value = color.theme
    else:
        value = None
    return {"type": color.type, "value": value, "tint": color.tint}


def _collect_differences(
    expected: Any,
    actual: Any,
    path: str,
    differences: list[str],
) -> None:
    if type(expected) is not type(actual):
        differences.append(
            f"{path}: type {type(expected).__name__} != {type(actual).__name__} "
            f"({expected!r} != {actual!r})"
        )
        return
    if isinstance(expected, dict):
        expected_keys = list(expected)
        actual_keys = list(actual)
        if expected_keys != actual_keys:
            differences.append(f"{path}: keys {expected_keys!r} != {actual_keys!r}")
        for key in expected_keys:
            if key not in actual:
                continue
            _collect_differences(
                expected[key],
                actual[key],
                f"{path}.{key}",
                differences,
            )
        return
    if isinstance(expected, list):
        if len(expected) != len(actual):
            differences.append(f"{path}: length {len(expected)} != {len(actual)}")
        for index, (expected_item, actual_item) in enumerate(zip(expected, actual)):
            _collect_differences(
                expected_item,
                actual_item,
                f"{path}[{index}]",
                differences,
            )
        return
    if expected != actual:
        differences.append(f"{path}: {expected!r} != {actual!r}")
