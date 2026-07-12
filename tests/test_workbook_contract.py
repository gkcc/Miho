from __future__ import annotations

import csv
import json
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

import pytest
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from hsr_endgame_exporter.exporters import write_all_outputs as write_hsr_outputs
from tests.workbook_contract import (
    FormulaPolicy,
    assert_workbooks_equal,
    compare_workbooks,
    workbook_semantics,
)
from zzz_endgame_exporter.exporters import write_outputs as write_zzz_outputs


FIXTURES = Path(__file__).parent / "fixtures" / "workbook_contract"
FORMULA_TEXT = '=HYPERLINK("https://invalid.example","external text")'

HSR_SHEETS = [
    "overview",
    "latest_usage_cn",
    "top_teams_latest",
    "phase_index",
    "character_usage_long",
    "character_usage_phase_latest",
    "histograph_usage_long",
    "team_rank_raw",
    "team_rank_dedup_ordered",
    "team_rank_dedup_unordered",
    "name_map",
    "name_map_unresolved",
    "prydwen_tier_current",
    "prydwen_tier_history",
    "prydwen_tier_changelog",
    "prydwen_tier_changelog_history",
    "prydwen_tier_usage_trend",
    "prydwen_tier_charts",
]

ZZZ_SHEETS = [
    "phase_index",
    "character_usage_long",
    "character_usage_phase_latest",
    "team_rank_raw",
    "team_rank_dedup_unordered",
    "name_map",
    "name_map_unresolved",
    "prydwen_tier_current",
    "prydwen_tier_history",
    "prydwen_tier_changelog",
    "prydwen_tier_changelog_history",
    "prydwen_tier_usage_trend",
]


@pytest.mark.parametrize(
    ("game", "writer"),
    [("hsr", lambda root: _write_hsr_oracle(root)), ("zzz", lambda root: _write_zzz_oracle(root))],
)
def test_python_writer_matches_semantic_oracle(
    tmp_path: Path,
    game: str,
    writer: object,
) -> None:
    actual = writer(tmp_path / game)  # type: ignore[operator]
    expected = _oracle_path(game)
    assert actual.name == expected.name
    assert_workbooks_equal(expected, actual)


@pytest.mark.parametrize(
    ("game", "writer"),
    [("hsr", lambda root: _write_hsr_oracle(root)), ("zzz", lambda root: _write_zzz_oracle(root))],
)
def test_rust_writer_matches_semantic_oracle_without_formulas(
    tmp_path: Path,
    game: str,
    writer: object,
) -> None:
    csv_root = tmp_path / f"{game}-csv"
    writer(csv_root)  # type: ignore[operator]
    actual = tmp_path / "rust" / _oracle_path(game).name
    _run_rust_writer(game, csv_root, actual)

    assert_workbooks_equal(
        _oracle_path(game),
        actual,
        formula_policy=FormulaPolicy.EXTERNAL_TEXT,
    )
    actual_semantics = workbook_semantics(actual)
    assert sum(sheet["formula_count"] for sheet in actual_semantics["sheets"]) == 0


def test_rust_writer_keeps_blank_formats_and_mixed_explicit_types(tmp_path: Path) -> None:
    hsr_root = tmp_path / "hsr-csv"
    for sheet in HSR_SHEETS:
        _write_csv(hsr_root / f"{sheet}.csv", ["text"], [])
    _write_csv(
        hsr_root / "character_usage_long.csv",
        ["phase_ver", "app_rate", "sample"],
        [["1.0", "", ""]],
    )
    _write_csv(
        hsr_root / "prydwen_tier_current.csv",
        ["special_rating"],
        [["8.5"], ["E6"]],
    )
    hsr_output = tmp_path / "hsr" / "hsr_endgame_dataset.xlsx"
    _run_rust_writer("hsr", hsr_root, hsr_output)
    hsr = load_workbook(hsr_output)
    try:
        sheet = hsr["character_usage_long"]
        assert (sheet["A2"].value, sheet["A2"].data_type) == ("1.0", "s")
        assert (sheet["B2"].value, sheet["B2"].data_type, sheet["B2"].number_format) == (
            None,
            "n",
            "0.00",
        )
        assert (sheet["C2"].value, sheet["C2"].data_type, sheet["C2"].number_format) == (
            None,
            "n",
            "0.00",
        )
        tier = hsr["prydwen_tier_current"]
        assert (tier["A2"].value, tier["A2"].data_type) == (8.5, "n")
        assert (tier["A3"].value, tier["A3"].data_type) == ("E6", "s")
    finally:
        hsr.close()

    zzz_root = tmp_path / "zzz-csv"
    for sheet in ZZZ_SHEETS:
        _write_csv(zzz_root / f"{sheet}.csv", ["text"], [])
    _write_csv(zzz_root / "character_usage_long.csv", ["rarity"], [["5"], ["S"]])
    _write_csv(
        zzz_root / "name_map.csv",
        ["release_order", "needs_manual_check"],
        [["10", "1"]],
    )
    zzz_output = tmp_path / "zzz" / "zzz_endgame_dataset.xlsx"
    _run_rust_writer("zzz", zzz_root, zzz_output)
    zzz = load_workbook(zzz_output)
    try:
        usage = zzz["character_usage_long"]
        assert (usage["A2"].value, usage["A2"].data_type) == (5, "n")
        assert (usage["A3"].value, usage["A3"].data_type) == ("S", "s")
        names = zzz["name_map"]
        assert (names["A2"].value, names["A2"].data_type) == ("10", "s")
        assert (names["B2"].value, names["B2"].data_type) == ("1", "s")
    finally:
        zzz.close()


def test_oracle_sheet_order_and_layout_contract() -> None:
    hsr = workbook_semantics(_oracle_path("hsr"))
    zzz = workbook_semantics(_oracle_path("zzz"))

    assert hsr["file_name"] == "hsr_endgame_dataset.xlsx"
    assert zzz["file_name"] == "zzz_endgame_dataset.xlsx"
    assert hsr["sheet_order"] == HSR_SHEETS
    assert zzz["sheet_order"] == ZZZ_SHEETS
    assert len(hsr["sheets"]) == 18
    assert len(zzz["sheets"]) == 12
    assert all(sheet["dimensions"].startswith("A1:") for sheet in hsr["sheets"])
    assert all(sheet["dimensions"].startswith("A1:") for sheet in zzz["sheets"])
    assert all(len(sheet["headers"]) > 0 for sheet in hsr["sheets"])
    assert all(len(sheet["headers"]) > 0 for sheet in zzz["sheets"])


def test_hsr_and_zzz_formatting_policies_are_frozen() -> None:
    hsr = workbook_semantics(_oracle_path("hsr"))
    zzz = workbook_semantics(_oracle_path("zzz"))
    hsr_sheets = {sheet["title"]: sheet for sheet in hsr["sheets"]}

    assert hsr["active_sheet"] == "overview"
    assert hsr_sheets["overview"]["freeze_panes"] is None
    assert all(
        sheet["freeze_panes"] == "A2"
        for sheet in hsr["sheets"]
        if sheet["title"] != "overview"
    )
    assert all(sheet["auto_filter"] == sheet["dimensions"] for sheet in hsr["sheets"])
    assert all(sheet["show_grid_lines"] is False for sheet in hsr["sheets"])
    assert all(width is not None for sheet in hsr["sheets"] for width in sheet["column_widths"].values())

    for title in ("overview", "latest_usage_cn", "top_teams_latest"):
        for style in hsr_sheets[title]["header_styles"]:
            assert style["fill"]["fill_type"] == "solid"
            assert style["fill"]["foreground"]["value"] == "E8F3F1"
            assert style["font"]["bold"] is True
            assert style["font"]["color"]["value"] == "1F2933"
            assert style["alignment"]["horizontal"] == "center"
            assert style["alignment"]["vertical"] == "center"

    for style in hsr_sheets["phase_index"]["header_styles"]:
        assert style["fill"]["fill_type"] == "solid"
        assert style["fill"]["foreground"]["value"] == "263238"
        assert style["font"]["bold"] is True
        assert style["font"]["color"]["value"] == "FFFFFF"

    for sheet in [*hsr["sheets"], *zzz["sheets"]]:
        for style in sheet["header_styles"]:
            assert style["border"] == {
                "left": "thin",
                "right": "thin",
                "top": "thin",
                "bottom": "thin",
                "diagonal_up": False,
                "diagonal_down": False,
            }

    usage_cells = {cell["coordinate"]: cell for cell in hsr_sheets["character_usage_long"]["cells"]}
    team_cells = {cell["coordinate"]: cell for cell in hsr_sheets["team_rank_raw"]["cells"]}
    assert usage_cells["P2"]["number_format"] == "0.00"
    assert usage_cells["V2"]["number_format"] == "0.00"
    assert team_cells["J2"]["number_format"] == "0.00"

    assert zzz["active_sheet"] == "phase_index"
    assert all(sheet["freeze_panes"] is None for sheet in zzz["sheets"])
    assert all(sheet["auto_filter"] is None for sheet in zzz["sheets"])
    assert all(sheet["show_grid_lines"] is True for sheet in zzz["sheets"])
    assert all(width is None for sheet in zzz["sheets"] for width in sheet["column_widths"].values())
    for style in zzz["sheets"][0]["header_styles"]:
        assert style["fill"]["fill_type"] is None
        assert style["font"]["bold"] is True
        assert style["alignment"]["horizontal"] == "center"
        assert style["alignment"]["vertical"] == "top"
    zzz_usage = {
        cell["coordinate"]: cell
        for cell in next(sheet for sheet in zzz["sheets"] if sheet["title"] == "character_usage_long")["cells"]
    }
    assert zzz_usage["P2"]["number_format"] == "General"


@pytest.mark.parametrize("game", ["hsr", "zzz"])
def test_external_formula_injection_baseline_and_safe_normalization(
    tmp_path: Path,
    game: str,
) -> None:
    oracle = _oracle_path(game)
    strict = workbook_semantics(oracle)
    changelog = next(
        sheet for sheet in strict["sheets"] if sheet["title"] == "prydwen_tier_changelog"
    )
    formula_cell = next(cell for cell in changelog["cells"] if cell["coordinate"] == "D2")
    assert formula_cell == {
        "coordinate": "D2",
        "value": FORMULA_TEXT,
        "data_type": "f",
        "number_format": "General",
    }
    assert changelog["formula_count"] == 1
    assert changelog["formula_cells"] == ["D2"]
    assert sum(sheet["formula_count"] for sheet in strict["sheets"]) == 1

    safe = tmp_path / oracle.name
    shutil.copyfile(oracle, safe)
    workbook = load_workbook(safe, data_only=False)
    cell = workbook["prydwen_tier_changelog"]["D2"]
    cell.value = FORMULA_TEXT
    cell.data_type = "s"
    workbook.save(safe)
    workbook.close()

    strict_differences = compare_workbooks(oracle, safe)
    assert any("formula_count" in difference for difference in strict_differences)
    assert any("data_type" in difference for difference in strict_differences)
    assert_workbooks_equal(
        oracle,
        safe,
        formula_policy=FormulaPolicy.EXTERNAL_TEXT,
    )


def test_fixture_manifest_documents_comparison_boundary() -> None:
    manifest = json.loads((FIXTURES / "contract.json").read_text(encoding="utf-8"))
    assert manifest["schema_version"] == 1
    assert manifest["ignored"] == [
        "ZIP member ordering and bytes",
        "ZIP member timestamps",
        "openpyxl style_id",
    ]
    assert manifest["normalized"] == [
        "RGB alpha channel",
        "inherited Calibri 11 font metadata",
        "unused solid-fill background",
        "default header border color",
        "grouped column dimension ranges",
    ]
    for game, sheets in (("hsr", HSR_SHEETS), ("zzz", ZZZ_SHEETS)):
        entry = manifest["oracles"][game]
        assert entry["file_name"] == _oracle_path(game).name
        assert entry["sheet_order"] == sheets
        assert entry["formula_injection_baseline"] == {
            "sheet": "prydwen_tier_changelog",
            "cell": "D2",
            "external_text": FORMULA_TEXT,
            "python_data_type": "f",
            "python_formula_count": 1,
            "rust_policy": "write as literal text",
        }


def _oracle_path(game: str) -> Path:
    return FIXTURES / game / f"{game}_endgame_dataset.xlsx"


def _write_csv(path: Path, headers: list[str], rows: list[list[str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle, lineterminator="\r\n")
        writer.writerow(headers)
        writer.writerows(rows)


def _run_rust_writer(game: str, csv_root: Path, output: Path) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    subprocess.run(
        [
            "cargo",
            "run",
            "--quiet",
            "-p",
            "miho-core",
            "--example",
            "workbook_contract",
            "--",
            game,
            str(csv_root),
            str(output),
        ],
        cwd=ROOT,
        check=True,
    )


def _write_hsr_oracle(out_dir: Path) -> Path:
    phase = {
        "snapshot_id": "fixture-1",
        "collect_date": "2026-07-12",
        "mode": "moc",
        "mode_cn": "混沌回忆",
        "phase_ver": "1.0",
        "phase_name": "脱敏阶段",
        "start_date": "2026-07-01",
        "end_date": "2026-07-14",
        "source": "fixture",
        "source_path": "fixture-1/",
        "has_chars": True,
        "has_comps": True,
        "has_histograph": True,
        "note": "最小脱敏样本",
    }
    usage = {
        "snapshot_id": "fixture-1",
        "collect_date": "2026-07-12",
        "mode": "moc",
        "mode_cn": "混沌回忆",
        "sub_mode": "all",
        "sub_mode_cn": "全部",
        "phase_ver": "1.0",
        "phase_name": "脱敏阶段",
        "start_date": "2026-07-01",
        "end_date": "2026-07-14",
        "character_slug": "agent-alpha",
        "character_name_en": "Agent Alpha",
        "character_name_cn": "代理甲",
        "role": "damage",
        "rarity": 5,
        "app_rate": 12.5,
        "app_rate_e0": 8.75,
        "avg_round": 3.25,
        "std_dev_round": 1.5,
        "q1_round": 2.0,
        "cons_avg": 0.5,
        "sample": 123,
        "sample_app_flat": 45,
        "source_kind": "hf_chars",
        "source_file": "fixture/builds.json",
        "source_url": "https://invalid.example/builds",
        "quality_flag": "ok",
    }
    histograph = {
        "snapshot_id": "fixture-1",
        "collect_date": "2026-07-12",
        "mode": "moc",
        "mode_cn": "混沌回忆",
        "character_slug": "agent-alpha",
        "character_name_en": "Agent Alpha",
        "character_name_cn": "代理甲",
        "usage_value": 12.5,
        "source_file": "fixture/histograph.json",
        "note": "脱敏直方图",
    }
    team = {
        "snapshot_id": "fixture-1",
        "collect_date": "2026-07-12",
        "mode": "moc",
        "mode_cn": "混沌回忆",
        "sub_mode": "all",
        "sub_mode_cn": "全部",
        "phase_ver": "1.0",
        "phase_name": "脱敏阶段",
        "scope": "all",
        "rank": 1,
        "comp_name": "脱敏队伍",
        "char_1_slug": "agent-alpha",
        "char_2_slug": "agent-beta",
        "char_3_slug": "agent-gamma",
        "char_4_slug": "agent-delta",
        "char_1_name_cn": "代理甲",
        "char_2_name_cn": "代理乙",
        "char_3_name_cn": "代理丙",
        "char_4_name_cn": "代理丁",
        "app_rate": 9.5,
        "avg_round": 4.25,
        "whale_count": 2,
        "app_flat": 19,
        "uses": 11,
        "source_kind": "hf_comps",
        "source_file": "fixture/teams.json",
        "source_url": "https://invalid.example/teams",
        "raw_index": 1,
        "raw_json": '{"redacted":true}',
    }
    tier = _hsr_tier_row("2026-07-12")
    tier_history = _hsr_tier_row("2026-07-01")
    trend = {
        "tier_snapshot_id": "2026-07-12",
        "tier_updated_date": "2026-07-12",
        "tier_mode": "moc",
        "tier_mode_cn": "混沌回忆",
        "character_slug": "agent-alpha",
        "character_name_en": "Agent Alpha",
        "character_name_cn": "代理甲",
        "prydwen_role": "Damage",
        "role_group": "damage",
        "role_group_cn": "输出",
        "tier": "T1",
        "rating": 9,
        "tags": "tag-a",
        "marks": "mark-a",
        "collect_date": "2026-07-12",
        "phase_ver": "1.0",
        "phase_name": "脱敏阶段",
        "app_rate": 12.5,
        "avg_round": 3.25,
        "quality_flag": "ok",
        "icon_url": "https://invalid.example/icon.png",
    }
    out_dir.mkdir(parents=True, exist_ok=True)
    write_hsr_outputs(
        out_dir,
        phase_rows=[phase],
        character_rows=[usage],
        histograph_rows=[histograph],
        team_raw_rows=[team],
        name_map_rows=[_hsr_name_row("agent-alpha", "0")],
        name_map_unresolved_rows=[_hsr_name_row("agent-beta", "1")],
        prydwen_tier_current_rows=[tier],
        prydwen_tier_history_rows=[tier_history],
        prydwen_tier_changelog_rows=[
            {
                "changelog_date": "2026-07-12",
                "source_url": "https://invalid.example/changelog",
                "character_slugs": "agent-alpha",
                "text": FORMULA_TEXT,
            }
        ],
        prydwen_tier_changelog_history_rows=[
            {
                "changelog_date": "2026-07-01",
                "source_url": "https://invalid.example/changelog-old",
                "character_slugs": "agent-alpha",
                "text": "历史文本",
            }
        ],
        prydwen_tier_usage_trend_rows=[trend],
        prydwen_tier_chart_rows=[
            {
                "tier_mode": "moc",
                "tier_mode_cn": "混沌回忆",
                "role_group": "damage",
                "role_group_cn": "输出",
                "chart_file": "charts/moc-damage.svg",
                "series_count": 1,
                "point_count": 2,
            }
        ],
        warnings=["脱敏警告"],
    )
    return out_dir / "hsr_endgame_dataset.xlsx"


def _write_zzz_oracle(out_dir: Path) -> Path:
    phase = {
        "snapshot_id": "fixture-1",
        "collect_date": "2026-07-12",
        "mode": "sd",
        "mode_cn": "式舆防卫",
        "phase_ver": "1.0",
        "phase_name": "脱敏阶段",
        "start_date": "2026-07-01",
        "end_date": "2026-07-14",
        "source": "fixture",
        "source_path": "fixture-1/",
        "has_chars": True,
        "has_comps": True,
        "note": "最小脱敏样本",
    }
    usage = {
        "snapshot_id": "fixture-1",
        "collect_date": "2026-07-12",
        "mode": "sd",
        "mode_cn": "式舆防卫",
        "sub_mode": "all",
        "sub_mode_cn": "全部",
        "phase_ver": "1.0",
        "phase_name": "脱敏阶段",
        "start_date": "2026-07-01",
        "end_date": "2026-07-14",
        "character_slug": "agent-alpha",
        "character_name_en": "Agent Alpha",
        "character_name_cn": "代理甲",
        "role": "attack",
        "rarity": 5,
        "app_rate": 23.5,
        "avg_score": 32123.5,
        "sample": 123,
        "sample_players": 100,
        "cons_avg": 0.5,
        "char_level": 60,
        "w_engine_level": 60,
        "core_skill": 7,
        "source_kind": "hf_builds",
        "source_file": "fixture/builds.json",
        "source_url": "https://invalid.example/builds",
        "quality_flag": "ok",
    }
    team = {
        "snapshot_id": "fixture-1",
        "collect_date": "2026-07-12",
        "mode": "sd",
        "mode_cn": "式舆防卫",
        "sub_mode": "all",
        "sub_mode_cn": "全部",
        "phase_ver": "1.0",
        "phase_name": "脱敏阶段",
        "scope": "all",
        "rank": 1,
        "char_1_slug": "agent-alpha",
        "char_2_slug": "agent-beta",
        "char_3_slug": "agent-gamma",
        "bangboo_slug": "bangboo-alpha",
        "char_1_name_cn": "代理甲",
        "char_2_name_cn": "代理乙",
        "char_3_name_cn": "代理丙",
        "bangboo_name_cn": "邦布甲",
        "app_rate": 11.25,
        "avg_score": 32000.5,
        "avg_score_m1": 33000.5,
        "source_kind": "hf_comps",
        "source_file": "fixture/teams.json",
        "source_url": "https://invalid.example/teams",
        "raw_index": 1,
        "raw_json": '{"redacted":true}',
    }
    tier = _zzz_tier_row("2026-07-12")
    trend = {
        **tier,
        "collect_date": "2026-07-12",
        "phase_ver": "1.0",
        "phase_name": "脱敏阶段",
        "app_rate": 23.5,
        "avg_score": 32123.5,
        "quality_flag": "ok",
    }
    out_dir.mkdir(parents=True, exist_ok=True)
    write_zzz_outputs(
        out_dir,
        phase_rows=[phase],
        usage_rows=[usage],
        team_rows=[team],
        name_rows=[_zzz_name_row("agent-alpha", "0")],
        unresolved_rows=[_zzz_name_row("agent-beta", "1")],
        tier_rows=[tier],
        tier_history_rows=[_zzz_tier_row("2026-07-01")],
        changelog_rows=[
            {
                "changelog_date": "2026-07-12",
                "source_url": "https://invalid.example/changelog",
                "character_slugs": "agent-alpha",
                "text": FORMULA_TEXT,
            }
        ],
        changelog_history_rows=[
            {
                "changelog_date": "2026-07-01",
                "source_url": "https://invalid.example/changelog-old",
                "character_slugs": "agent-alpha",
                "text": "历史文本",
            }
        ],
        trend_rows=[trend],
        warnings=["脱敏警告"],
    )
    return out_dir / "zzz_endgame_dataset.xlsx"


def _hsr_tier_row(snapshot: str) -> dict[str, object]:
    return {
        "tier_snapshot_id": snapshot,
        "fetched_at": f"{snapshot}T00:00:00Z",
        "tier_updated_at": snapshot,
        "tier_updated_date": snapshot,
        "tier_mode": "moc",
        "tier_mode_cn": "混沌回忆",
        "character_slug": "agent-alpha",
        "character_name_en": "Agent Alpha",
        "character_name_cn": "代理甲",
        "prydwen_category": "Damage",
        "prydwen_role": "Damage",
        "role_group": "damage",
        "role_group_cn": "输出",
        "tier": "T1",
        "rating": 9,
        "special_rating": 8.5,
        "tags": "tag-a",
        "marks": "mark-a",
        "is_new": True,
        "default_role": "damage",
        "element": "Imaginary",
        "path": "Hunt",
        "rarity": 5,
        "icon_url": "https://invalid.example/icon.png",
        "source_url": "https://invalid.example/tier",
    }


def _zzz_tier_row(snapshot: str) -> dict[str, object]:
    return {
        "tier_snapshot_id": snapshot,
        "fetched_at": f"{snapshot}T00:00:00Z",
        "tier_updated_at": snapshot,
        "tier_updated_date": snapshot,
        "tier_mode": "sd",
        "tier_mode_cn": "式舆防卫",
        "character_slug": "agent-alpha",
        "character_name_en": "Agent Alpha",
        "character_name_cn": "代理甲",
        "prydwen_category": "CritDPS",
        "prydwen_role": "Attack",
        "role_group": "crit_dps",
        "role_group_cn": "直伤主C",
        "tier": "T1",
        "rating": 9,
        "tags": "tag-a",
        "marks": "mark-a",
        "is_new": True,
        "element": "Ether",
        "element_cn": "以太",
        "style": "Attack",
        "style_cn": "强攻",
        "faction": "Redacted",
        "rarity": 5,
        "icon_url": "https://invalid.example/icon.png",
        "source_url": "https://invalid.example/tier",
    }


def _hsr_name_row(slug: str, unresolved: str) -> dict[str, object]:
    return {
        "character_slug": slug,
        "character_name_en": slug.replace("-", " ").title(),
        "character_name_cn": "待确认" if unresolved == "1" else "代理甲",
        "source": "fixture",
        "needs_manual_check": unresolved,
        "aliases": f"{slug}-alias",
    }


def _zzz_name_row(slug: str, unresolved: str) -> dict[str, object]:
    return {
        **_hsr_name_row(slug, unresolved),
        "kind": "agent",
        "release_order": "10",
    }


def _regenerate_oracles() -> None:
    with tempfile.TemporaryDirectory(prefix="miho-workbook-oracle-") as temporary:
        temporary_root = Path(temporary)
        for game, writer in (("hsr", _write_hsr_oracle), ("zzz", _write_zzz_oracle)):
            generated = writer(temporary_root / game)
            target = _oracle_path(game)
            target.parent.mkdir(parents=True, exist_ok=True)
            shutil.copyfile(generated, target)


if __name__ == "__main__":
    _regenerate_oracles()
